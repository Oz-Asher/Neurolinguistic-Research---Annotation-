
import os
import sys
import subprocess

subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
os.system('cls' if os.name == 'nt' else 'clear')  # Clearing the terminal

from collections import OrderedDict
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter


# This code creates and edits the excel of the disagreements

# -------------------- PART 1: Format Headers --------------------
def format_headers(ws):
    """
    Adjusts column widths and freezes the header row.
    Returns a list of tuples (col_index, header_value).
    """
    headers = [(i + 1, cell.value) for i, cell in enumerate(ws[1]) if isinstance(cell.value, str)]
    for col_idx, header in headers:
        ws.column_dimensions[get_column_letter(col_idx)].width = len(header) + 2
    ws.freeze_panes = "A2"
    return headers


# -------------------- PART 2: Color Header Titles --------------------
def color_header_cells(ws, headers):
    """
    Colors header titles based on their theme group.
    """

    theme_colors = OrderedDict()
    for _, header in headers:
        prefix = header.split()[0]
        if prefix not in theme_colors:
            theme_colors[prefix] = None
    colors = ["000FFF", "00FF00"]  # Blue and Green
    for i, theme in enumerate(theme_colors):
        theme_colors[theme] = colors[i % 2]

    for col_idx, header in headers:
        prefix = header.split()[0]
        color = theme_colors.get(prefix, "FFFFFF")  # Default white if missing
        cell = ws[f"{get_column_letter(col_idx)}1"]
        cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

# -------------------- PART 3: Highlight Disagreements --------------------
def highlight_m_cells(ws):
    """
    Highlights cells that contain '@M' in yellow, and cleans them.
    """
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if isinstance(cell.value, str) and "@M" in cell.value:
                cell.fill = yellow_fill
                cell.value = cell.value.replace('@M', '').replace('nan', '').replace('None', '')

# -------------------- PART 4: Controller Function --------------------
def excel_editing(combined, directory):
    """
    Main function to save the DataFrame and apply formatting & coloring.
    """
    path = os.path.join(directory, "Disagreements.xlsx")

    combined.to_excel(path, index=False) # Save the given DataFrame to an Excel file at the specified path.

    wb = load_workbook(path)
    ws = wb.active

    headers = format_headers(ws)
    color_header_cells(ws, headers)
    highlight_m_cells(ws)

    wb.save(path)
    print(f"Excel saved at: {path}")

